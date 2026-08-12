"""Seed Medipost group companies, directors, PI members, and Kawari asset list."""
from __future__ import annotations

import json
import re
import uuid
from collections import defaultdict
from pathlib import Path

import openpyxl

XLSX = Path(r"c:\Users\devth\Downloads\Medipost - Companies Layout (1).xlsx")
OUT = Path(__file__).resolve().parents[1] / "data"

PARENT_ID = "3a842763-e5f2-43f7-87c2-2a5429738fb3"
PARENT_ZOHO = "7351644000001876029"
KAWARI_ID = "b1111111-1111-4111-8111-111111111147"
KAWARI_ZOHO = "7351644000003979001"
MEDILOG_ID = "b1111111-1111-4111-8111-111111111151"
MEDILOG_ZOHO = "7351644000003987001"
PHARMACY_ID = "b1111111-1111-4111-8111-111111111152"
PHARMACY_ZOHO = "7351644000003987002"

# Stable branch ids
KAWARI_MIDRAND = "c4444444-4444-4444-8444-444444444401"
KAWARI_CPT = "c4444444-4444-4444-8444-444444444402"
MEDILOG_HO = "c4444444-4444-4444-8444-444444444403"
PHARMACY_GEZINA = "c4444444-4444-4444-8444-444444444404"
MEDIPOST_HO = "1fc97cdd-ff84-4afa-8bca-3840eafb5847"  # existing Head Office

NS = uuid.UUID("a1111111-1111-4111-8111-111111111001")

# Contact emails matched to directors (from Summary contacts)
DIRECTOR_EMAILS = {
    "Louis Scheepers": "louis@medipost.co.za",
    "Emmerentia Frederika Myburgh": "rentia@medipost.co.za",
    "Martha Sophia Joubert": None,
    "Willem Adolph Joubert": None,
    "Urvashi Maganlal": None,
    "Selapeng Chriscentia Maledimo": None,
    "Khulekani Mcwallace Dlamini": None,
    "Bulelwa Promise Mamabolo": None,
    "Ramoroesi Victor Jonas Ramathesele": None,
    "Mamedupi Matsipa": None,
    "Mziwandile Noel Guliwe": None,
}

# Company memberships for directors
DIRECTOR_ACCOUNTS = {
    "Louis Scheepers": [PARENT_ID, MEDILOG_ID, PHARMACY_ID],
    "Willem Adolph Joubert": [PARENT_ID, PHARMACY_ID],
    "Martha Sophia Joubert": [PARENT_ID, PHARMACY_ID],
    "Urvashi Maganlal": [PARENT_ID, PHARMACY_ID],
    "Selapeng Chriscentia Maledimo": [PARENT_ID, PHARMACY_ID],
    "Khulekani Mcwallace Dlamini": [PARENT_ID, PHARMACY_ID],
    "Bulelwa Promise Mamabolo": [MEDILOG_ID, PHARMACY_ID],
    "Ramoroesi Victor Jonas Ramathesele": [MEDILOG_ID, PHARMACY_ID],
    "Emmerentia Frederika Myburgh": [PHARMACY_ID],
    "Mamedupi Matsipa": [KAWARI_ID],
    "Mziwandile Noel Guliwe": [KAWARI_ID],
}

# Portal login home = primary company (NOT parent holding). Branch-scoped.
DIRECTOR_HOME = {
    "Louis Scheepers": (MEDILOG_ID, MEDILOG_HO),
    "Willem Adolph Joubert": (PARENT_ID, MEDIPOST_HO),
    "Martha Sophia Joubert": (PHARMACY_ID, PHARMACY_GEZINA),
    "Urvashi Maganlal": (PARENT_ID, MEDIPOST_HO),
    "Selapeng Chriscentia Maledimo": (PARENT_ID, MEDIPOST_HO),
    "Khulekani Mcwallace Dlamini": (PARENT_ID, MEDIPOST_HO),
    "Bulelwa Promise Mamabolo": (MEDILOG_ID, MEDILOG_HO),
    "Ramoroesi Victor Jonas Ramathesele": (MEDILOG_ID, MEDILOG_HO),
    "Emmerentia Frederika Myburgh": (PHARMACY_ID, PHARMACY_GEZINA),
    "Mamedupi Matsipa": (KAWARI_ID, KAWARI_MIDRAND),
    "Mziwandile Noel Guliwe": (KAWARI_ID, KAWARI_MIDRAND),
}


def sql_str(v) -> str:
    if v is None:
        return "null"
    return "'" + str(v).replace("'", "''") + "'"


def sql_num(v):
    if v is None or v == "":
        return "null"
    try:
        return str(float(v))
    except Exception:
        return "null"


def director_id(name: str) -> str:
    return str(uuid.uuid5(NS, f"director:{name.strip().lower()}"))


def invite_email(name: str, known: str | None) -> str:
    if known:
        return known.lower()
    slug = re.sub(r"[^a-z0-9]+", ".", name.lower()).strip(".")
    return f"{slug}@directors.aegis.local"


def parse_directors_from_summary(wb) -> set[str]:
    ws = wb["Summary - 20250810"]
    names: set[str] = set()
    for r in range(4, 14):
        raw = ws.cell(r, 4).value
        if not raw:
            continue
        for part in re.split(r"[\n/;]+", str(raw)):
            n = part.strip()
            # strip trailing whitespace junk
            n = re.sub(r"\s+", " ", n).strip()
            if len(n) >= 5 and not n.startswith("."):
                names.add(n)
    return names


def parse_pi_members(wb) -> list[dict]:
    ws = wb["PI Members List"]
    rows: list[dict] = []
    cover_month = None
    cover_year = 2025
    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        if c1 is None:
            continue
        s = str(c1).strip()
        if s.startswith("PROFESSIONAL") or s.startswith("Member") or s.startswith("Year 20"):
            m = re.search(r"20\d{2}", s)
            if m:
                cover_year = int(m.group(0))
            continue
        if hasattr(c1, "year"):
            cover_month = c1.date() if hasattr(c1, "date") else c1
            cover_year = cover_month.year
            continue
        entity = ws.cell(r, 2).value
        if not entity:
            continue
        id_raw = ws.cell(r, 3).value
        id_number = None
        if id_raw is not None:
            if isinstance(id_raw, float):
                id_number = str(int(id_raw))
            else:
                id_number = str(id_raw).strip()
        account_id = PHARMACY_ID if "Medipost" in str(entity) else KAWARI_ID
        rows.append(
            {
                "account_id": account_id,
                "policy_number": "PI313688",
                "full_name": s,
                "entity_name": str(entity).strip(),
                "id_number": id_number,
                "email": (str(ws.cell(r, 4).value).strip() if ws.cell(r, 4).value else None),
                "phone": (str(ws.cell(r, 5).value).strip() if ws.cell(r, 5).value else None),
                "council_number": (str(ws.cell(r, 6).value).strip() if ws.cell(r, 6).value else None),
                "vat_amount": ws.cell(r, 7).value,
                "premium": ws.cell(r, 8).value,
                "payment_status": (str(ws.cell(r, 9).value).strip() if ws.cell(r, 9).value else None),
                "comments": (str(ws.cell(r, 10).value).strip() if ws.cell(r, 10).value else None),
                "cover_year": cover_year,
                "cover_month": cover_month.isoformat() if cover_month else f"{cover_year}-01-01",
                "source_row": r,
            }
        )
    return rows


def parse_assets(wb) -> list[dict]:
    ws = wb["Asset List - 20260808"]
    assets: list[dict] = []
    for r in range(3, ws.max_row + 1):
        cat = ws.cell(r, 1).value
        desc = ws.cell(r, 8).value
        section = ws.cell(r, 9).value
        if not desc or not cat:
            continue
        cat_s = str(cat).strip()
        if cat_s in {"Row Labels", "Grand Total", "First Loss", "Category"} or not re.match(
            r"^[A-Z]\d+", cat_s
        ):
            continue
        branch_code = str(ws.cell(r, 4).value or "").strip()
        alt = ws.cell(r, 7).value
        asset_num = ws.cell(r, 6).value
        cost = ws.cell(r, 11).value
        loc = ws.cell(r, 5).value
        branch_id = KAWARI_CPT if branch_code == "1003" else KAWARI_MIDRAND
        section_s = str(section).strip() if section else "Miscellaneous"
        category = (
            "Electronic Equipment"
            if "Electronic" in section_s
            else "Contents"
            if "Content" in section_s or "Office" in section_s
            else "Building"
            if section_s == "Fire"
            else "Miscellaneous"
        )
        tag = str(alt).strip() if alt else f"KW-{asset_num}"
        assets.append(
            {
                "id": str(uuid.uuid5(NS, f"asset:{tag}")),
                "name": str(desc).strip()[:200],
                "asset_tag": tag[:80],
                "category": category,
                "insurance_section": section_s[:80],
                "unit_cost": float(cost) if cost not in (None, "") else 0,
                "branch_id": branch_id,
                "branch": "Cape Town" if branch_code == "1003" else "Midrand",
                "location_code": str(loc).strip() if loc else None,
                "asset_number": str(asset_num) if asset_num is not None else None,
                "class_code": cat_s,
            }
        )
    return assets


def build_sql(pi_rows: list[dict], assets: list[dict], director_names: set[str]) -> list[Path]:
    OUT.mkdir(exist_ok=True)
    parts: list[Path] = []

    # --- accounts + branches + directors ---
    core = []
    core.append("-- Medipost group accounts, branches, directors")
    core.append("BEGIN;")
    # Update parent registration
    core.append(
        f"UPDATE portal_accounts SET registration_number = '2014/089301/07', name = 'Medipost FTRR&I' WHERE id = '{PARENT_ID}';"
    )
    # Kawari already exists — refresh reg
    core.append(
        f"""
INSERT INTO portal_accounts (id, name, parent_account_id, zoho_account_id, industry, registration_number, phone, aegis_status)
VALUES ('{KAWARI_ID}', 'Kawari Wholesalers (Pty) Ltd', '{PARENT_ID}', '{KAWARI_ZOHO}', 'Wholesale', '2004/015737/07', '+27 12 426 4001', 'active')
ON CONFLICT (id) DO UPDATE SET parent_account_id = EXCLUDED.parent_account_id, zoho_account_id = EXCLUDED.zoho_account_id,
  registration_number = EXCLUDED.registration_number, aegis_status = 'active';
""".strip()
    )
    core.append(
        f"""
INSERT INTO portal_accounts (id, name, parent_account_id, zoho_account_id, industry, registration_number, phone, aegis_status)
VALUES ('{MEDILOG_ID}', 'Medilogistics FTRR&I', '{PARENT_ID}', '{MEDILOG_ZOHO}', 'Logistics', '2011/129984/07', '+27 12 426 4000', 'active')
ON CONFLICT (id) DO UPDATE SET parent_account_id = EXCLUDED.parent_account_id, zoho_account_id = EXCLUDED.zoho_account_id,
  registration_number = EXCLUDED.registration_number, name = EXCLUDED.name, aegis_status = 'active';
""".strip()
    )
    core.append(
        f"""
INSERT INTO portal_accounts (id, name, parent_account_id, zoho_account_id, industry, registration_number, phone, aegis_status)
VALUES ('{PHARMACY_ID}', 'Medipost Pharmacy (HH Durrheim)', '{PARENT_ID}', '{PHARMACY_ZOHO}', 'Pharmacy', '1997/011099/07', '+27 12 426 4000', 'active')
ON CONFLICT (id) DO UPDATE SET parent_account_id = EXCLUDED.parent_account_id, zoho_account_id = EXCLUDED.zoho_account_id,
  registration_number = EXCLUDED.registration_number, name = EXCLUDED.name, aegis_status = 'active';
""".strip()
    )

    branches = [
        (KAWARI_MIDRAND, KAWARI_ID, "Midrand", "Crescent Corporate Park, Unit A1, Block A, Technohub 59 Roan, Midrand, Gauteng, 1682", -25.996, 28.128),
        (KAWARI_CPT, KAWARI_ID, "Cape Town", "Parow / Cape Town operations", -33.8985, 18.5902),
        (MEDILOG_HO, MEDILOG_ID, "Head Office", "Gezina City Centre, 593 Nico Smith Street, Gezina, Pretoria, Gauteng, 0002", -25.7205, 28.2055),
        (PHARMACY_GEZINA, PHARMACY_ID, "Gezina", "Gezina City Centre, 593 Nico Smith Street, Gezina, Pretoria, Gauteng, 0002", -25.7205, 28.2055),
    ]
    for bid, aid, name, addr, lat, lng in branches:
        core.append(
            f"""
INSERT INTO portal_branches (id, account_id, name, address, latitude, longitude)
VALUES ('{bid}', '{aid}', {sql_str(name)}, {sql_str(addr)}, {lat}, {lng})
ON CONFLICT (id) DO UPDATE SET name = EXCLUDED.name, address = EXCLUDED.address, latitude = EXCLUDED.latitude, longitude = EXCLUDED.longitude;
""".strip()
        )

    # Directors
    for name in sorted(director_names | set(DIRECTOR_ACCOUNTS)):
        did = director_id(name)
        email = invite_email(name, DIRECTOR_EMAILS.get(name))
        core.append(
            f"""
INSERT INTO portal_directors (id, full_name, email)
VALUES ('{did}', {sql_str(name)}, {sql_str(email)})
ON CONFLICT (id) DO UPDATE SET full_name = EXCLUDED.full_name, email = COALESCE(EXCLUDED.email, portal_directors.email), updated_at = now();
""".strip()
        )
        for aid in DIRECTOR_ACCOUNTS.get(name, [PARENT_ID]):
            core.append(
                f"""
INSERT INTO portal_director_accounts (director_id, account_id, title, is_primary)
VALUES ('{did}', '{aid}', 'Director', true)
ON CONFLICT (director_id, account_id) DO UPDATE SET title = EXCLUDED.title;
""".strip()
            )
        home_account, home_branch = DIRECTOR_HOME.get(name, (PARENT_ID, MEDIPOST_HO))
        # Invite for branch-scoped login on their company (not parent holding view)
        core.append(
            f"""
INSERT INTO portal_invites (account_id, email, role, branch_id, director_id)
SELECT '{home_account}', {sql_str(email)}, 'member', '{home_branch}', '{did}'
WHERE NOT EXISTS (
  SELECT 1 FROM portal_invites WHERE lower(email) = lower({sql_str(email)}) AND accepted_at IS NULL
);
""".strip()
        )

    # Placeholder PI policy if missing
    core.append(
        f"""
INSERT INTO portal_policies (
  id, account_id, zoho_policy_id, policy_number, status, premium, inception_date, renewal_date,
  insurer, product_line, frequency, covered_items
) VALUES (
  'c2222222-2222-4222-8222-222222222313',
  '{PHARMACY_ID}',
  'pi313688-rsum',
  'PI313688',
  'Active',
  1665,
  '2025-04-01',
  '2026-04-01',
  'RSUM',
  'Professional Indemnity (Pharmacists)',
  'Monthly',
  '[]'::jsonb
)
ON CONFLICT (id) DO UPDATE SET account_id = EXCLUDED.account_id, premium = EXCLUDED.premium;
""".strip()
    )
    core.append("COMMIT;")
    p0 = OUT / "seed-medipost-group-core.sql"
    p0.write_text("\n".join(core), encoding="utf-8")
    parts.append(p0)

    # --- PI members ---
    pi_sql = ["BEGIN;", "DELETE FROM portal_pi_members WHERE policy_number = 'PI313688';"]
    for row in pi_rows:
        pid = str(uuid.uuid5(NS, f"pi:{row['policy_number']}:{row['id_number']}:{row['full_name']}:{row['cover_month']}"))
        pi_sql.append(
            f"""
INSERT INTO portal_pi_members (
  id, account_id, policy_id, policy_number, full_name, entity_name, id_number, email, phone,
  council_number, vat_amount, premium, payment_status, cover_year, cover_month, comments, source_row
) VALUES (
  '{pid}',
  '{row['account_id']}',
  'c2222222-2222-4222-8222-222222222313',
  {sql_str(row['policy_number'])},
  {sql_str(row['full_name'])},
  {sql_str(row['entity_name'])},
  {sql_str(row['id_number'])},
  {sql_str(row['email'])},
  {sql_str(row['phone'])},
  {sql_str(row['council_number'])},
  {sql_num(row['vat_amount'])},
  {sql_num(row['premium'])},
  {sql_str(row['payment_status'])},
  {row['cover_year'] if row['cover_year'] else 'null'},
  {sql_str(row['cover_month'])}::date,
  {sql_str(row['comments'])},
  {row['source_row']}
)
ON CONFLICT (id) DO UPDATE SET
  premium = EXCLUDED.premium,
  payment_status = EXCLUDED.payment_status,
  vat_amount = EXCLUDED.vat_amount,
  comments = EXCLUDED.comments,
  updated_at = now();
""".strip()
        )
    pi_sql.append("COMMIT;")
    p1 = OUT / "seed-medipost-pi-members.sql"
    p1.write_text("\n".join(pi_sql), encoding="utf-8")
    parts.append(p1)

    # --- Assets in chunks ---
    chunk_size = 150
    for i in range(0, len(assets), chunk_size):
        chunk = assets[i : i + chunk_size]
        lines = ["BEGIN;"]
        for a in chunk:
            zoho_fields = json.dumps(
                {
                    "asset_list_code": a["class_code"],
                    "asset_number": a["asset_number"],
                    "location_code": a["location_code"],
                    "source": "Asset List - 20260808",
                    "owner_account": "Kawari",
                }
            )
            lines.append(
                f"""
INSERT INTO portal_risk_items (
  id, account_id, zoho_risk_id, name, category, insurance_section, unit_cost, repair_cost,
  record_date, insurance_status, asset_tag, description, branch_id, branch, zoho_fields,
  latitude, longitude
) VALUES (
  '{a['id']}',
  '{KAWARI_ID}',
  {sql_str('asset-' + a['asset_tag'])},
  {sql_str(a['name'])},
  {sql_str(a['category'])},
  {sql_str(a['insurance_section'])},
  {a['unit_cost']},
  0,
  '2026-02-01',
  'Insured with us',
  {sql_str(a['asset_tag'])},
  {sql_str(a['name'])},
  '{a['branch_id']}',
  {sql_str(a['branch'])},
  {sql_str(zoho_fields)}::jsonb,
  CASE WHEN '{a['branch_id']}' = '{KAWARI_CPT}' THEN -33.8985 ELSE -25.996 END,
  CASE WHEN '{a['branch_id']}' = '{KAWARI_CPT}' THEN 18.5902 ELSE 28.128 END
)
ON CONFLICT (id) DO UPDATE SET
  name = EXCLUDED.name,
  category = EXCLUDED.category,
  insurance_section = EXCLUDED.insurance_section,
  unit_cost = EXCLUDED.unit_cost,
  asset_tag = EXCLUDED.asset_tag,
  branch_id = EXCLUDED.branch_id,
  branch = EXCLUDED.branch,
  zoho_fields = EXCLUDED.zoho_fields,
  insurance_status = EXCLUDED.insurance_status,
  updated_at = now();
""".strip()
            )
            # Also upsert by asset_tag if a different id already exists
            lines.append(
                f"""
UPDATE portal_risk_items SET
  account_id = '{KAWARI_ID}',
  name = {sql_str(a['name'])},
  category = {sql_str(a['category'])},
  insurance_section = {sql_str(a['insurance_section'])},
  unit_cost = {a['unit_cost']},
  branch_id = '{a['branch_id']}',
  branch = {sql_str(a['branch'])},
  zoho_fields = {sql_str(zoho_fields)}::jsonb,
  insurance_status = 'Insured with us',
  updated_at = now()
WHERE account_id IN ('{KAWARI_ID}', '{PARENT_ID}')
  AND asset_tag = {sql_str(a['asset_tag'])}
  AND id <> '{a['id']}';
""".strip()
            )
        lines.append("COMMIT;")
        path = OUT / f"seed-kawari-assets-{i // chunk_size:02d}.sql"
        path.write_text("\n".join(lines), encoding="utf-8")
        parts.append(path)

    meta = {
        "directors": sorted(director_names | set(DIRECTOR_ACCOUNTS)),
        "pi_count": len(pi_rows),
        "asset_count": len(assets),
        "accounts": {
            "medipost": PARENT_ID,
            "kawari": KAWARI_ID,
            "medilogistics": MEDILOG_ID,
            "pharmacy": PHARMACY_ID,
        },
    }
    (OUT / "seed-medipost-group-meta.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")
    return parts


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    directors = parse_directors_from_summary(wb)
    pi = parse_pi_members(wb)
    assets = parse_assets(wb)
    paths = build_sql(pi, assets, directors)
    print(f"directors={len(directors)} pi={len(pi)} assets={len(assets)}")
    for p in paths:
        print(p.name, p.stat().st_size)


if __name__ == "__main__":
    main()
